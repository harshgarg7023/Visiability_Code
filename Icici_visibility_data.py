import os
import sys
import json
import argparse
import collections

try:
    import psycopg2
    import psycopg2.extras
except ImportError:
    sys.exit("Missing dependency. Run: pip install psycopg2-binary")


# =====================================================================
# CREDENTIALS — fill these in directly.
#
# WARNING: once real values are here, do NOT commit this file to git,
# do NOT upload/paste it anywhere (including back into this chat), and
# do NOT share it with anyone. If it's ever exposed, change the Aurora
# password immediately afterward.
# =====================================================================


# Populate os.environ so the rest of the script (which reads from env vars)
# works unchanged.
os.environ.setdefault("AURORA_HOST", AURORA_HOST)
os.environ.setdefault("AURORA_PORT", AURORA_PORT)
os.environ.setdefault("AURORA_DB", AURORA_DB)
os.environ.setdefault("AURORA_USER", AURORA_USER)
os.environ.setdefault("AURORA_PASSWORD", AURORA_PASSWORD)
os.environ.setdefault("AURORA_SSLMODE", AURORA_SSLMODE)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

 
  
 

# =====================================================================
# CONFIG — adjust here if column/value names differ on other insurers
# =====================================================================
 
# subproduct_id -> friendly sub-product name (confirmed from real data)
SUBPRODUCT_ID_MAP = {
    1: "Two Wheeler (TW)",
    2: "Private Car (PC)",
    9: "PCV (Passenger Comm. Vehicle)",
    10: "GCV (Goods Carrying Vehicle)",
    17: "misD (Miscellaneous Vehicle)",
}
SUB_PRODUCT_ORDER = [
    "Two Wheeler (TW)",
    "Private Car (PC)",
    "PCV (Passenger Comm. Vehicle)",
    "GCV (Goods Carrying Vehicle)",
    "misD (Miscellaneous Vehicle)",
    "Unknown",
]
SUB_PRODUCT_SHORT = {
    "Two Wheeler (TW)": "TW",
    "Private Car (PC)": "PC",
    "PCV (Passenger Comm. Vehicle)": "PCV",
    "GCV (Goods Carrying Vehicle)": "GCV",
    "misD (Miscellaneous Vehicle)": "MISC",
    "Unknown": "Others",
}
 
# segment column -> friendly Plan Type name (confirmed real distinct values)
SEGMENT_PLAN_MAP = {
    "Third Party": "Third Party",
    "Comprehensive": "Comprehensive",
    "OD Only": "Own Damage (SAOD)",
}
PLAN_ORDER = ["Third Party", "Comprehensive", "Own Damage (SAOD)", "Unclassified"]
PLAN_SHORT = {
    "Third Party": "TP",
    "Comprehensive": "COMP",
    "Own Damage (SAOD)": "SAOD",
    "Unclassified": "UNCL",
}
 
# Status value on the response row that means "a valid quote was generated"
SUCCESS_STATUS_VALUE = "Success"
 
# Column name on the response table holding the failure reason text, used
# for the Top 10 Errors sheet. Update this if the real column differs
# (e.g. 'error_message', 'failure_reason', 'derived_error_message').
FLAT_ERROR_MSG_COL_CANDIDATES = ["error_message", "failure_reason", "derived_error_message"]
 
# Error Details sheet: header order. The 7 most useful fields for quick
# triage (Segment, Plan, RegNo, Insurer Reference, Make, Model, Error) come
# first; the remaining fields keep their original relative order after that.
# Only columns with a confident real-column mapping are populated below
# (see fetch_error_detail_rows) - anything without a confirmed source is
# left BLANK rather than guessed, per instruction.
ERROR_DETAIL_HEADERS = [
    "Segment", "Plan", "RegNo", "Insurer Reference", "Make", "Model", "Error",
    "Transaction Type", "product", "RTO Name", "Registration Date", "Proposal Date",
    "C.Y. NCB", "P.Y. NCB", "Claim(Yes/No) in PYP", "PYP End Date",
    "Current year start Date", "Ownership serial number", "PYP details", "PYP type",
    "Add-on in PYP", "Add-on covers opted in proposal", "PAYU-", "PAYU-Plan",
]
 
# Max rows pulled into the Error Details sheet (failed/no-response rows can
# be numerous; cap to keep the workbook a reasonable size). Raise if needed.
ERROR_DETAIL_ROW_LIMIT = 10000
 
ADDON_FLAG_COLUMNS = {
    "is_cng": "CNG",
    "is_rti": "RTI",
    "is_consumables": "Consumables",
    "is_engine_protect": "Engine Protect",
    "is_pa_cover_owner_driver": "PA Cover (Owner-Driver)",
    "is_pa_cover_waiver": "PA Cover Waiver",
}
 
 
# =====================================================================
# DB CONNECTION
# =====================================================================
 
def get_connection():
    required = ["AURORA_HOST", "AURORA_DB", "AURORA_USER", "AURORA_PASSWORD"]
    missing = [v for v in required if not os.environ.get(v)]
    if missing:
        sys.exit(
            f"Missing required environment variables: {', '.join(missing)}\n"
            f"Create a '.env' file next to this script with these keys set "
            f"(see the SETUP section at the top of this file for the format)."
        )
    try:
        return psycopg2.connect(
            host=os.environ["AURORA_HOST"],
            port=os.environ.get("AURORA_PORT", "5432"),
            dbname=os.environ["AURORA_DB"],
            user=os.environ["AURORA_USER"],
            password=os.environ["AURORA_PASSWORD"],
            sslmode=os.environ.get("AURORA_SSLMODE", "require"),
            connect_timeout=15,
        )
    except psycopg2.OperationalError as e:
        sys.exit(
            "Could not connect to Aurora.\n"
            f"Details: {e}\n\n"
            "Common causes:\n"
            "  - Aurora is inside a private VPC and not reachable from your current "
            "network (needs VPN / bastion host / running from an EC2 instance in the same VPC)\n"
            "  - Security group on the RDS instance doesn't allow your IP on port 5432\n"
            "  - Wrong host, port, database name, username, or password in .env\n"
            "  - sslmode mismatch (try setting AURORA_SSLMODE=disable in .env if your "
            "instance doesn't enforce SSL, though 'require' is safer if it does)"
        )
 
 
def discover_insurer_tables(cur):
    """Find every <x>_raw_requests / <x>_raw_responses pair in public schema."""
    cur.execute("""
        SELECT table_name FROM information_schema.tables
        WHERE table_schema = 'public'
        ORDER BY table_name;
    """)
    tables = {r[0] for r in cur.fetchall()}
    pairs = []
    for t in tables:
        if t.endswith("_raw_requests"):
            insurer = t[: -len("_raw_requests")]
            resp_table = f"{insurer}_raw_responses"
            if resp_table in tables:
                pairs.append((insurer.upper(), t, resp_table))
    return pairs
 
 
def find_column(cur, table_name, candidates):
    cur.execute("""
        SELECT column_name FROM information_schema.columns
        WHERE table_schema = 'public' AND table_name = %s;
    """, (table_name,))
    existing = {r[0] for r in cur.fetchall()}
    for c in candidates:
        if c in existing:
            return c
    return None
 
 
# =====================================================================
# SQL-SIDE AGGREGATION
# =====================================================================
 
def fetch_segment_and_plan_counts(cur, req_table, resp_table, insurer_code,
                                   ic_col_req, ic_col_resp, created_col,
                                   start_date, end_date):
    """
    Runs the SubProduct x Plan Type aggregation directly in SQL, using
    COUNT(DISTINCT ... CASE WHEN ...) so a correlation_id with multiple
    response rows is only ever counted once, and counts as a success if
    ANY of its responses succeeded (mirrors the confirmed-working query
    style, extended to also break down by segment/plan type together).
 
    Returns a list of dicts: subproduct_id, segment, requests, responses, quotes
    """
    date_clause = ""
    if created_col and (start_date or end_date):
        if start_date:
            date_clause += f' AND req."{created_col}" >= %s'
        if end_date:
            date_clause += f' AND req."{created_col}" <= %s'
 
    ic_join_clause = f'AND rsp."{ic_col_resp}" = %s' if ic_col_resp else ""
 
    query = f"""
        SELECT
            COALESCE(req.subproduct_id, 0) AS subproduct_id,
            COALESCE(req.segment, 'Unknown') AS segment,
            COUNT(DISTINCT req.metadata_correlation_id) AS requests,
            COUNT(DISTINCT rsp.metadata_correlation_id) AS responses,
            COUNT(DISTINCT CASE
                WHEN rsp.derived_status = %s
                THEN rsp.metadata_correlation_id
            END) AS quotes
        FROM "{req_table}" req
        LEFT JOIN "{resp_table}" rsp
            ON req.metadata_correlation_id = rsp.metadata_correlation_id
            {ic_join_clause}
        WHERE req."{ic_col_req}" = %s
        {date_clause}
        GROUP BY req.subproduct_id, req.segment
        ORDER BY requests DESC;
    """
    exec_params = [SUCCESS_STATUS_VALUE]
    if ic_col_resp:
        exec_params.append(insurer_code)
    exec_params.append(insurer_code)
    if created_col and (start_date or end_date):
        if start_date:
            exec_params.append(start_date)
        if end_date:
            exec_params.append(end_date)
 
    cur.execute(query, exec_params)
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, row)) for row in cur.fetchall()]
 
 
def fetch_error_reasons_by_segment(cur, req_table, resp_table, insurer_code, error_col,
                                    ic_col_req, ic_col_resp, created_col, start_date, end_date):
    """
    Same idea as fetch_top_errors, but grouped ALSO by req.subproduct_id and
    req.segment, so each Sub-Product x Plan Type combo gets its own ranked
    failure-reason breakdown instead of one combined list across everything.
    """
    date_clause = ""
    if created_col and (start_date or end_date):
        if start_date:
            date_clause += f' AND req."{created_col}" >= %s'
        if end_date:
            date_clause += f' AND req."{created_col}" <= %s'
 
    ic_join_clause = f'AND rsp."{ic_col_resp}" = %s' if ic_col_resp else ""
    error_select = f'COALESCE(NULLIF(TRIM(rsp."{error_col}"), \'\'), \'Unknown Failure\')' if error_col else "'Unknown Failure'"
 
    query = f"""
        SELECT subproduct_id, segment, reason, SUM(cnt) AS total_count FROM (
            SELECT req.subproduct_id, req.segment, {error_select} AS reason, COUNT(*) AS cnt
            FROM "{req_table}" req
            JOIN "{resp_table}" rsp
                ON req.metadata_correlation_id = rsp.metadata_correlation_id
                {ic_join_clause}
            WHERE req."{ic_col_req}" = %s
              AND rsp.derived_status <> %s
              {date_clause}
            GROUP BY req.subproduct_id, req.segment, reason
 
            UNION ALL
 
            SELECT req.subproduct_id, req.segment, 'NO RESPONSE RECEIVED' AS reason, COUNT(*) AS cnt
            FROM "{req_table}" req
            WHERE req."{ic_col_req}" = %s
              AND NOT EXISTS (
                  SELECT 1 FROM "{resp_table}" rsp2
                  WHERE rsp2.metadata_correlation_id = req.metadata_correlation_id
              )
              {date_clause}
            GROUP BY req.subproduct_id, req.segment
        ) sub
        GROUP BY subproduct_id, segment, reason
        ORDER BY subproduct_id, segment, total_count DESC;
    """
    exec_params = []
    if ic_col_resp:
        exec_params.append(insurer_code)
    exec_params.append(insurer_code)
    exec_params.append(SUCCESS_STATUS_VALUE)
    if created_col and (start_date or end_date):
        if start_date:
            exec_params.append(start_date)
        if end_date:
            exec_params.append(end_date)
    exec_params.append(insurer_code)
    if created_col and (start_date or end_date):
        if start_date:
            exec_params.append(start_date)
        if end_date:
            exec_params.append(end_date)
 
    cur.execute(query, exec_params)
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, row)) for row in cur.fetchall()]
 
 
def fetch_error_detail_rows(cur, req_table, resp_table, insurer_code, error_col,
                             ic_col_req, ic_col_resp, created_col, start_date, end_date,
                             limit=ERROR_DETAIL_ROW_LIMIT):
    """
    Pulls one row per FAILED or NO-RESPONSE request, with only the flat
    columns we have a confirmed mapping for. Anything without a confirmed
    real column stays absent here and is left blank in the sheet - never
    guessed.
    """
    date_clause = ""
    if created_col and (start_date or end_date):
        if start_date:
            date_clause += f' AND req."{created_col}" >= %s'
        if end_date:
            date_clause += f' AND req."{created_col}" <= %s'
 
    ic_join_clause = f'AND rsp."{ic_col_resp}" = %s' if ic_col_resp else ""
    error_select = f'rsp."{error_col}"' if error_col else "NULL"
 
    addon_cols_sql = ", ".join(f'req."{c}"' for c in ADDON_FLAG_COLUMNS)
 
    # Only include rows that actually HAVE an error message - not every
    # failed/no-response row. Rows with no response at all (no error_message
    # possible) and failures where the error text is blank/NULL are excluded
    # entirely, per instruction: only show what we actually have.
    if error_col:
        error_filter = f'AND rsp."{error_col}" IS NOT NULL AND TRIM(rsp."{error_col}") <> \'\''
    else:
        # No error-message column was found at all, so nothing can qualify.
        error_filter = "AND FALSE"
 
    query = f"""
        SELECT
            req.metadata_correlation_id AS correlation_id,
            {error_select} AS error_remark,
            req.business_type,
            req.make_code,
            req.model_code,
            req.subproduct_id,
            req.registration_number,
            req.first_reg_date,
            req."{created_col}" AS proposal_date,
            req.ncb_percent,
            req.claim_on_prev_policy,
            req.prev_policy_end_date,
            req.policy_start_date,
            req.prev_policy_type,
            req.segment,
            {addon_cols_sql}
        FROM "{req_table}" req
        JOIN "{resp_table}" rsp
            ON req.metadata_correlation_id = rsp.metadata_correlation_id
            {ic_join_clause}
        WHERE req."{ic_col_req}" = %s
          AND rsp.derived_status IS DISTINCT FROM %s
          {error_filter}
          {date_clause}
        ORDER BY req."{created_col if created_col else 'metadata_correlation_id'}" DESC
        LIMIT {int(limit)};
    """
    exec_params = []
    if ic_col_resp:
        exec_params.append(insurer_code)
    exec_params.append(insurer_code)
    exec_params.append(SUCCESS_STATUS_VALUE)
    if created_col and (start_date or end_date):
        if start_date:
            exec_params.append(start_date)
        if end_date:
            exec_params.append(end_date)
 
    cur.execute(query, exec_params)
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, row)) for row in cur.fetchall()]
 
 
def shape_error_detail_row(raw):
    """Maps a raw DB row (from fetch_error_detail_rows) into the exact
    ERROR_DETAIL_HEADERS column order. Anything without a confirmed source
    column is left as None (blank cell) - never guessed."""
    sp = SUBPRODUCT_ID_MAP.get(raw.get("subproduct_id"), "")
    claim = raw.get("claim_on_prev_policy")
    claim_display = "Yes" if claim is True else ("No" if claim is False else None)
 
    addons_opted = [label for col, label in ADDON_FLAG_COLUMNS.items() if raw.get(col) is True]
    addons_display = ", ".join(addons_opted) if addons_opted else None
 
    return [
        sp,                                                                    # Segment
        raw.get("segment"),                                               # Plan (Current year requirement / Plan Type)
        raw.get("registration_number"),                       # RegNo
        raw.get("correlation_id"),                    # Insurer Reference (Co-relation id)
        raw.get("make_code"),                            # Make
        raw.get("model_code"),                            # Model
        raw.get("error_remark"),                       # Error
        raw.get("business_type"),                       # Transaction Type
        sp,                                                 # product
        None,                                                # RTO Name (no confirmed source)
        raw.get("first_reg_date"),                             # Registration Date
        raw.get("proposal_date"),                               # Proposal Date
        None,                                                     # C.Y. NCB (no confirmed source)
        raw.get("ncb_percent"),                                   # P.Y. NCB
        claim_display,                                              # Claim(Yes/No) in PYP
        raw.get("prev_policy_end_date"),                             # PYP End Date
        raw.get("policy_start_date"),                                 # Current year start Date
        None,                                                          # Ownership serial number (no confirmed source)
        None,                                                           # PYP details (no confirmed source)
        raw.get("prev_policy_type"),                                     # PYP type
        None,                                                              # Add-on in PYP (no confirmed source)
        addons_display,                                                     # Add-on covers opted in proposal
        None,                                                                # PAYU- (no confirmed source)
        None,                                                                 # PAYU-Plan (no confirmed source)
    ]
 
 
# =====================================================================
# AGGREGATION / SHAPING
# =====================================================================
 
def build_report_data(conn, insurer_filter=None, start_date=None, end_date=None):
    schema_cur = conn.cursor()
    data_cur = conn.cursor()
 
    print("Discovering insurer tables...", flush=True)
    pairs = discover_insurer_tables(schema_cur)
    print(f"Found {len(pairs)} insurer table pair(s): {[p[0] for p in pairs]}", flush=True)
 
    if insurer_filter:
        pairs = [p for p in pairs if p[0] == insurer_filter.upper()]
        if not pairs:
            sys.exit(f"No tables found for insurer '{insurer_filter}'.")
 
    sub_product_counts = collections.defaultdict(lambda: {"total": 0, "responses": 0, "quotes": 0})
    plan_counts = collections.defaultdict(lambda: {"total": 0, "responses": 0, "quotes": 0})
    error_counts_by_segment = collections.defaultdict(collections.Counter)  # key: (sp, plan)
    error_detail_rows = []
 
    for insurer, req_table, resp_table in pairs:
        print(f"\nProcessing insurer: {insurer}", flush=True)
 
        ic_col_req = find_column(schema_cur, req_table, ["ic"]) or "ic"
        ic_col_resp = find_column(schema_cur, resp_table, ["ic"])
        created_col = find_column(schema_cur, req_table, ["sqs_timestamp", "created_at", "createdat"])
        error_col = find_column(schema_cur, resp_table, FLAT_ERROR_MSG_COL_CANDIDATES)
 
        print("  Running segment/plan aggregation query...", flush=True)
        rows = fetch_segment_and_plan_counts(
            data_cur, req_table, resp_table, insurer,
            ic_col_req, ic_col_resp, created_col, start_date, end_date,
        )
        print(f"  -> {len(rows)} (sub-product, plan) group(s) found", flush=True)
 
        for row in rows:
            sp = SUBPRODUCT_ID_MAP.get(row["subproduct_id"], "Unknown")
            plan = SEGMENT_PLAN_MAP.get(row["segment"], "Unclassified")
 
            sub_product_counts[sp]["total"] += row["requests"]
            sub_product_counts[sp]["responses"] += row["responses"]
            sub_product_counts[sp]["quotes"] += row["quotes"]
 
            pkey = (sp, plan)
            plan_counts[pkey]["total"] += row["requests"]
            plan_counts[pkey]["responses"] += row["responses"]
            plan_counts[pkey]["quotes"] += row["quotes"]
 
        print("  Running failure reason aggregation query (per segment)...", flush=True)
        try:
            err_rows = fetch_error_reasons_by_segment(
                data_cur, req_table, resp_table, insurer, error_col,
                ic_col_req, ic_col_resp, created_col, start_date, end_date,
            )
            for er in err_rows:
                sp = SUBPRODUCT_ID_MAP.get(er["subproduct_id"], "Unknown")
                plan = SEGMENT_PLAN_MAP.get(er["segment"], "Unclassified")
                error_counts_by_segment[(sp, plan)][er["reason"]] += er["total_count"]
        except Exception as e:
            print(f"  WARNING: failure-reason query failed ({e}); "
                  f"Error Summary sheet may be incomplete for {insurer}. "
                  f"Check FLAT_ERROR_MSG_COL_CANDIDATES matches your real column name.", flush=True)
 
        print("  Fetching Error Details rows (failed / no-response requests)...", flush=True)
        try:
            raw_detail_rows = fetch_error_detail_rows(
                data_cur, req_table, resp_table, insurer, error_col,
                ic_col_req, ic_col_resp, created_col, start_date, end_date,
            )
            print(f"  -> {len(raw_detail_rows)} error-detail row(s) fetched "
                  f"(capped at {ERROR_DETAIL_ROW_LIMIT})", flush=True)
            for raw in raw_detail_rows:
                error_detail_rows.append(shape_error_detail_row(raw))
        except Exception as e:
            print(f"  WARNING: Error Details query failed ({e}); "
                  f"Error Details sheet will be left blank for {insurer}.", flush=True)
 
    schema_cur.close()
    data_cur.close()
    return sub_product_counts, plan_counts, error_counts_by_segment, error_detail_rows
 
 
# =====================================================================
# XLSX OUTPUT — styling matches the approved Probus sample exactly
# =====================================================================
 
NAVY = "1F4E78"
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
KPI_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
TITLE_FONT = Font(name="Arial", bold=True, size=14, color=NAVY)
SECTION_FONT = Font(name="Calibri", bold=True, size=11)
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11)
BODY_FONT = Font(name="Calibri", size=11)
BODY_BOLD_FONT = Font(name="Calibri", bold=True, size=11)
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
LABEL_FONT = Font(name="Calibri", size=11, color=NAVY)
NOTE_FONT = Font(name="Calibri", italic=True, size=9, color="808080")
thin = Side(style="thin", color="B7B7B7")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
 
ERR_TITLE_FONT = Font(name="Arial", bold=True, size=14, color=NAVY)
ERR_HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
ERR_BODY_FONT = Font(name="Arial", size=11)
 
 
def style_header_row(ws, row, first_col, last_col, font=HEADER_FONT):
    for c in range(first_col, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = font
        cell.alignment = CENTER
        cell.border = BORDER
 
 
def style_data_row(ws, row, first_col, last_col, bold=False, body_font=BODY_FONT, bold_font=BODY_BOLD_FONT):
    for c in range(first_col, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = BORDER
        cell.alignment = CENTER
        cell.font = bold_font if bold else body_font
 
 
def autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
 
 
def build_workbook(insurer_label, date_label, sub_product_counts, plan_counts, error_counts_by_segment, out_path,
                    error_detail_rows=None):
    wb = openpyxl.Workbook()
 
    total_total = sum(c["total"] for c in sub_product_counts.values())
    total_resp = sum(c["responses"] for c in sub_product_counts.values())
    total_quotes = sum(c["quotes"] for c in sub_product_counts.values())
    overall_response_pct = (total_resp / total_total) if total_total else 0
    overall_visibility_pct = (total_quotes / total_resp) if total_resp else 0
 
    # ---------------- Sheet 1: Overview ----------------
    ws1 = wb.active
    ws1.title = "Overview"
    ws1.sheet_view.showGridLines = False
 
    ws1["B2"] = f"Probus Motor Insurer Visibility Report - {insurer_label}"
    ws1["B2"].font = TITLE_FONT
 
    ws1["B4"] = date_label
    ws1["B4"].font = SUBTITLE_FONT
 
    ws1["B6"] = "High Level Summary"
    ws1["B6"].font = SECTION_FONT
    ws1["B6"].alignment = Alignment(horizontal="left")
    ws1.merge_cells("B6:F6")
 
    hl_headers = ["Quote Requests", "Quote Responses", "Response %", "Valid Quotations", "Visbility %"]
    for i, h in enumerate(hl_headers):
        c = ws1.cell(row=7, column=2 + i, value=h)
        c.font = BODY_FONT
        c.alignment = CENTER
        c.border = BORDER
 
    hl_values = [total_total, total_resp, overall_response_pct, total_quotes, overall_visibility_pct]
    for i, v in enumerate(hl_values):
        col = 2 + i
        c = ws1.cell(row=8, column=col, value=v)
        c.font = BODY_BOLD_FONT
        c.alignment = CENTER
        c.border = BORDER
        if col in (4, 6):  # Response % / Visbility %
            c.number_format = "0.0%"
            c.fill = KPI_FILL
        else:
            c.number_format = "#,##0"
 
    ws1["B10"] = "* Remarks and ranking to be displayed here"
    ws1["B10"].font = NOTE_FONT
 
    # --- Sub-Product Wise Visibility ---
    ws1["B13"] = "Sub-Product Wise Visbility"
    ws1["B13"].font = SECTION_FONT
 
    sp_headers = ["Sub-Product", "Quote Responses", "Quotes Generated", "Visbility %", "Remarks"]
    for i, h in enumerate(sp_headers):
        ws1.cell(row=14, column=2 + i, value=h)
    style_header_row(ws1, 14, 2, 6)
 
    ordered_sps = [sp for sp in SUB_PRODUCT_ORDER if sp in sub_product_counts]
    r = 15
    for sp in ordered_sps:
        c = sub_product_counts[sp]
        responses = c["responses"]
        quotes = c["quotes"]
        valid_quote_pct = (quotes / responses) if responses else 0
        ws1.cell(row=r, column=2, value=SUB_PRODUCT_SHORT.get(sp, sp))
        ws1.cell(row=r, column=3, value=responses)
        ws1.cell(row=r, column=4, value=quotes)
        ws1.cell(row=r, column=5, value=valid_quote_pct)
        ws1.cell(row=r, column=5).number_format = "0.0%"
        style_data_row(ws1, r, 2, 6)
        ws1.cell(row=r, column=2).font = LABEL_FONT
        r += 1
    sp_block_end = r - 1
 
    # --- Plan Wise Visibility ---
    plan_section_row = sp_block_end + 3
    ws1.cell(row=plan_section_row, column=2, value="Plan Wise Visibility").font = SECTION_FONT
 
    pw_header_row = plan_section_row + 1
    pw_headers = ["Plan Type", "Quote Responses", "Quotes Generated", "Visbility %", "Remarks"]
    for i, h in enumerate(pw_headers):
        ws1.cell(row=pw_header_row, column=2 + i, value=h)
    style_header_row(ws1, pw_header_row, 2, 6)
 
    r = pw_header_row + 1
    for sp in ordered_sps:
        for plan in PLAN_ORDER:
            key = (sp, plan)
            if key not in plan_counts:
                continue
            c = plan_counts[key]
            if c["total"] == 0:
                continue
            valid_quote_pct = (c["quotes"] / c["responses"]) if c["responses"] else 0
            label = f"{SUB_PRODUCT_SHORT.get(sp, sp)}-{PLAN_SHORT.get(plan, plan)}"
            ws1.cell(row=r, column=2, value=label)
            ws1.cell(row=r, column=3, value=c["responses"])
            ws1.cell(row=r, column=4, value=c["quotes"])
            ws1.cell(row=r, column=5, value=valid_quote_pct)
            ws1.cell(row=r, column=5).number_format = "0.0%"
            style_data_row(ws1, r, 2, 6)
            ws1.cell(row=r, column=2).font = LABEL_FONT
            r += 1
 
    autofit(ws1, [3, 20, 20, 24, 23, 16])
 
    # ---------------- Sheet 2: Error Summary (per Sub-Product x Plan segment) ----------------
    ws2 = wb.create_sheet("Error Summary")
    ws2.sheet_view.showGridLines = False
    ws2["B2"] = f"Failure Analysis by Segment — {insurer_label}"
    ws2["B2"].font = ERR_TITLE_FONT
    ws2["B4"] = date_label
    ws2["B4"].font = SUBTITLE_FONT
 
    headers2 = ["Rank", "Failure Reason", "Count", "% of Segment Failure"]
    r = 6
    segment_keys = sorted(
        error_counts_by_segment.keys(),
        key=lambda k: f"{SUB_PRODUCT_SHORT.get(k[0], k[0])}-{PLAN_SHORT.get(k[1], k[1])}",
    )
    for sp, plan in segment_keys:
        label = f"{SUB_PRODUCT_SHORT.get(sp, sp)}-{PLAN_SHORT.get(plan, plan)}"
        ws2.cell(row=r, column=2, value=label).font = SECTION_FONT
        r += 1
 
        for i, h in enumerate(headers2):
            ws2.cell(row=r, column=2 + i, value=h)
        style_header_row(ws2, r, 2, 5, font=ERR_HEADER_FONT)
        r += 1
 
        top10 = error_counts_by_segment[(sp, plan)].most_common(10)
        total_failures = sum(count for _, count in top10)
        for rank, (reason, count) in enumerate(top10, start=1):
            display_reason = (reason[:250] + "…") if len(reason) > 250 else reason
            pct = (count / total_failures) if total_failures else 0
            ws2.cell(row=r, column=2, value=rank)
            ws2.cell(row=r, column=3, value=display_reason)
            ws2.cell(row=r, column=4, value=count)
            ws2.cell(row=r, column=5, value=pct)
            ws2.cell(row=r, column=5).number_format = "0.0%"
            style_data_row(ws2, r, 2, 5, body_font=ERR_BODY_FONT, bold_font=ERR_BODY_FONT)
            ws2.cell(row=r, column=3).alignment = LEFT
            # Auto-size row height for wrapped, multi-line failure reasons -
            # openpyxl doesn't auto-fit row height for wrapped text like
            # Excel sometimes does, so long error messages get visually
            # squished into a single line without this.
            chars_per_line = 60  # approx characters that fit in the 65-wide Failure Reason column
            est_lines = max(1, -(-len(display_reason) // chars_per_line))  # ceil division
            ws2.row_dimensions[r].height = max(15, est_lines * 15)
            r += 1
 
        r += 2  # blank rows between segments
 
    autofit(ws2, [3, 10, 65, 12, 20])
 
    # ---------------- Sheet 3: Error Details ----------------
    ws3 = wb.create_sheet("Error Details")
    ws3.sheet_view.showGridLines = False
 
    for i, h in enumerate(ERROR_DETAIL_HEADERS, start=1):
        ws3.cell(row=1, column=i, value=h)
    style_header_row(ws3, 1, 1, len(ERROR_DETAIL_HEADERS), font=ERR_HEADER_FONT)
 
    if error_detail_rows:
        for r_idx, detail_row in enumerate(error_detail_rows, start=2):
            for c_idx, value in enumerate(detail_row, start=1):
                cell = ws3.cell(row=r_idx, column=c_idx, value=value)
                cell.font = ERR_BODY_FONT
                cell.border = BORDER
                cell.alignment = LEFT if c_idx in (4, 7) else CENTER
 
    autofit(ws3, [16, 16, 16, 22, 12, 12, 45] + [16] * (len(ERROR_DETAIL_HEADERS) - 7))
 
    wb.save(out_path)
 
 
# =====================================================================
# MAIN
# =====================================================================
 
def main():
    parser = argparse.ArgumentParser(description="Generate Insurer Visibility Report")
    parser.add_argument("--insurer", help="Insurer code, e.g. ICICI (default: all insurers combined)")
    parser.add_argument("--start", help="Start date YYYY-MM-DD (filters request sqs_timestamp)")
    parser.add_argument("--end", help="End date YYYY-MM-DD (filters request sqs_timestamp)")
    parser.add_argument("--out", default="Probus Motor Insurer Visibility Report - ICICI.xlsx", help="Output file path")
    args = parser.parse_args()
 
    conn = get_connection()
    print("Connected to Aurora OK.", flush=True)
 
    sub_product_counts, plan_counts, error_counts_by_segment, error_detail_rows = build_report_data(
        conn, insurer_filter=args.insurer, start_date=args.start, end_date=args.end,
    )
    conn.close()
 
    if not sub_product_counts:
        sys.exit("No rows found for the given filters. Check --insurer / --start / --end.")
 
    label = args.insurer.upper() if args.insurer else "All Insurers"
 
    if args.start and args.end:
        date_label = f"{args.start} to {args.end}"
    elif args.start:
        date_label = f"From {args.start}"
    elif args.end:
        date_label = f"Through {args.end}"
    else:
        date_label = "All available data"
 
    try:
        build_workbook(label, date_label, sub_product_counts, plan_counts, error_counts_by_segment, args.out,
                        error_detail_rows=error_detail_rows)
        print(f"Report written to: {args.out}")
    except PermissionError:
        import datetime
        base, ext = os.path.splitext(args.out)
        fallback_out = f"{base}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
        print(f"WARNING: could not write to '{args.out}' (file may be open in Excel, "
              f"read-only, or locked by OneDrive). Retrying with a new filename...", flush=True)
        build_workbook(label, date_label, sub_product_counts, plan_counts, error_counts_by_segment, fallback_out,
                        error_detail_rows=error_detail_rows)
        print(f"Report written to: {fallback_out}")
 
 
if __name__ == "__main__":
    main()
 